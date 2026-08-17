"""processor.py'nin urettigi KPI ve tablo verilerini, yonetim raporu / e-posta
raporu formatina benzer sekilde TEK bir Excel sayfasinda ust uste akan bolumler
halinde yazar.

`openpyxl`, xlsx iceriginin XML/UTF-8 olarak yazilmasini garanti eder; bu yuzden
Turkce karakterler (İ, Ş, ç, ğ, ü, ö...) icin ekstra bir kodlama ayari gerekmez.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["create_excel_report"]

# --------------------------------------------------------------------------
# Bicimlendirme sabitleri
# --------------------------------------------------------------------------

TITLE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
TITLE_FONT = Font(bold=True, size=12, color="1F4E78")

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 60
ROWS_BETWEEN_SECTIONS = 2  # bolumler arasindaki bos satir sayisi
CHART_ROW_SPAN = 16  # grafigin kapladigi yaklasik satir sayisi (bir sonraki bolumle çakismamasi icin)

BAR_CHART_TITLE = "İterasyon Bazlı İş Büyüklüğü (SP)"

# Ust bolumdeki bar chart / buyukluk tablosunda gosterilecek 4 metrik ve Turkce etiketleri.
METRIC_COLUMNS: list[tuple[str, str]] = [
    ("committed_sp", "Taahhüt Edilen"),
    ("completed_sp", "Gerçekleşen"),
    ("out_of_plan_sp", "Plan Dışı"),
    ("total_completed_sp", "Toplam Tamamlanan"),
]

# processor.py tablolarindaki kolon adlarini rapor gereksinimindeki "(Sp)" etiketine cevirir.
PLANNED_COLUMN_RENAME = {
    "Hedeflenen Büyüklük": "Hedeflenen Büyüklük (Sp)",
    "Gerçekleşen Büyüklük": "Gerçekleşen Büyüklük (Sp)",
}
OUT_OF_PLAN_COLUMN_RENAME = {
    "Gerçekleşen Büyüklük": "Gerçekleşen Büyüklük (Sp)",
}


@dataclass
class _TableRange:
    header_row: int
    last_data_row: int
    next_row: int  # bolumden sonraki bosluk satirlari dahil, bir sonraki bolumun baslangic satiri


# --------------------------------------------------------------------------
# Yardimci fonksiyonlar
# --------------------------------------------------------------------------


def _build_bar_dataframe(iteration_history: list[tuple[str, dict]]) -> pd.DataFrame:
    """Son iterasyonlarin `summarize_metrics` ciktilarini İterasyon/4-metrik tablosuna cevirir."""
    rows = [
        {
            "İterasyon": label,
            **{turkish_label: summary.get(key, 0) for key, turkish_label in METRIC_COLUMNS},
        }
        for label, summary in iteration_history
    ]
    return pd.DataFrame(rows)


def _write_table(
    ws: Worksheet,
    start_row: int,
    title: str,
    df: pd.DataFrame,
    column_widths: dict[int, int],
) -> _TableRange:
    """`start_row`'dan itibaren bolum basligi + tablo basligi + veri satirlarini yazar."""
    span = max(len(df.columns), 1)

    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    if span > 1:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=span)

    header_row = start_row + 1
    numeric_cols = {column: pd.api.types.is_numeric_dtype(df[column]) for column in df.columns}

    for col_idx, column in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        column_widths[col_idx] = max(column_widths.get(col_idx, 0), len(str(column)))

    for row_offset, (_, record) in enumerate(df.iterrows(), start=1):
        row_idx = header_row + row_offset
        for col_idx, column in enumerate(df.columns, start=1):
            value = record[column]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_numeric = numeric_cols[column]
            cell.alignment = Alignment(horizontal="right" if is_numeric else "left", vertical="center")
            if is_numeric and isinstance(value, float):
                cell.number_format = "#,##0.00"

            column_widths[col_idx] = max(
                column_widths.get(col_idx, 0), len(str(value)) if value is not None else 0
            )

    last_data_row = header_row + len(df)
    next_row = last_data_row + 1 + ROWS_BETWEEN_SECTIONS
    return _TableRange(header_row=header_row, last_data_row=last_data_row, next_row=next_row)


def _write_bar_chart_section(
    ws: Worksheet,
    start_row: int,
    iteration_history: list[tuple[str, dict]],
    column_widths: dict[int, int],
) -> int:
    """İterasyon x (Taahhüt Edilen/Gerçekleşen/Plan Dışı/Toplam Tamamlanan) tablosunu ve
    bu veriye dayanan gomulu bir Excel bar chart'ini yazar."""
    df = _build_bar_dataframe(iteration_history)
    table_range = _write_table(ws, start_row, BAR_CHART_TITLE, df, column_widths)

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = BAR_CHART_TITLE
    chart.y_axis.title = "SP"
    chart.x_axis.title = "İterasyon"
    chart.height = 8
    chart.width = 16
    chart.style = 10

    data = Reference(
        ws,
        min_col=2,
        max_col=1 + len(METRIC_COLUMNS),
        min_row=table_range.header_row,
        max_row=table_range.last_data_row,
    )
    categories = Reference(
        ws, min_col=1, min_row=table_range.header_row + 1, max_row=table_range.last_data_row
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart_anchor_row = table_range.last_data_row + 1 + ROWS_BETWEEN_SECTIONS
    ws.add_chart(chart, f"A{chart_anchor_row}")

    return chart_anchor_row + CHART_ROW_SPAN + ROWS_BETWEEN_SECTIONS


def _apply_column_widths(ws: Worksheet, column_widths: dict[int, int]) -> None:
    for col_idx, max_len in column_widths.items():
        width = min(max(max_len + 2, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# --------------------------------------------------------------------------
# Genel API
# --------------------------------------------------------------------------


def create_excel_report(
    processed_data: dict,
    output_path: str | Path,
    target_month: str | None = None,
    iteration_history: list[tuple[str, dict]] | None = None,
) -> Path:
    """`process_sprint_report` ciktisini, yonetim raporu formatina benzer sekilde
    TEK bir Excel sayfasinda ust uste akan uc bolum halinde yazar:

        1) "İterasyon Bazlı İş Büyüklüğü (SP)" basligi + Taahhüt Edilen/Gerçekleşen/
           Plan Dışı/Toplam Tamamlanan kolonlu tablo ve buna dayanan gomulu bar chart
        2) "<ay> iterasyonunda planlanan iş listemiz ve statüleri:" basligi + tablo
        3) "<ay> iterasyonunda plan dışı iş listemiz ve statüleri:" basligi + tablo

    `processed_data`, `data`, `planned_issues`, `out_of_plan_issues`, `summary`,
    `kpis` ve `target_month` anahtarlarini iceren sozluktur (bkz.
    `processor.process_sprint_report`). `planned_issues`/`out_of_plan_issues` zaten
    `processor.process_sprint_report`'a verilen `target_month`'a gore filtrelenmis
    oldugundan burada tekrar filtrelenmez.

    Bolum 2 ve 3'un basligindaki ay etiketi oncelik sirasiyla belirlenir: bu fonksiyona
    verilen `target_month` parametresi verilirse o kullanilir, yoksa
    `processed_data["target_month"]` (processor.py'nin `created` tarihine gore belirledigi
    ay) kullanilir; ikisi de yoksa ay on eki olmadan genel ifadeler yazilir (orn.
    "Planlanan iş listemiz ve statüleri:").

    `iteration_history`, en eski once olacak sekilde `(iterasyon_adi, summary_dict)`
    ciftlerinden olusan bir listedir (orn. hedef ay Mayıs 2026 ise: [("Ocak 2026", ...),
    ("Şubat 2026", ...), ..., ("Mayıs 2026", ...)]) ve bar chart'ta karsilastirma icin
    kullanilir. Verilmezse, `processed_data["monthly_history"]` (bkz.
    `processor.process_sprint_report` / `build_yearly_monthly_history` - `created`
    tarihine gore, HEDEF AYIN (`target_month`) icinde bulundugu TAKVIM YILINDA
    Ocak'tan hedef aya kadar DAHIL olan aylar - "yil-basindan-bugune" karsilastirma;
    hedef aydan SONRAKI aylar veride olsa bile dahil edilmez) kullanilir; o da
    yoksa/bossa tek satirlik bir gecmis `target_month` (veya "Güncel İterasyon") ve
    `processed_data["summary"]` ile olusturulur.

    `output_path`'in ust klasoru (orn. `outputs/`) yoksa otomatik olusturulur.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if iteration_history is None:
        iteration_history = processed_data.get("monthly_history") or [
            (target_month or "Güncel İterasyon", processed_data["summary"])
        ]

    label = target_month or processed_data.get("target_month")
    prefix = f"{label} iterasyonunda " if label else ""
    planned_title = f"{prefix}planlanan iş listemiz ve statüleri:" if prefix else "Planlanan iş listemiz ve statüleri:"
    out_of_plan_title = f"{prefix}plan dışı iş listemiz ve statüleri:" if prefix else "Plan dışı iş listemiz ve statüleri:"

    planned_df = processed_data["planned_issues"].rename(columns=PLANNED_COLUMN_RENAME)
    out_of_plan_df = processed_data["out_of_plan_issues"].rename(columns=OUT_OF_PLAN_COLUMN_RENAME)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sprint Raporu"

    column_widths: dict[int, int] = {}
    row = 1
    row = _write_bar_chart_section(ws, row, iteration_history, column_widths)
    row = _write_table(ws, row, planned_title, planned_df, column_widths).next_row
    row = _write_table(ws, row, out_of_plan_title, out_of_plan_df, column_widths).next_row

    _apply_column_widths(ws, column_widths)

    try:
        wb.save(output_path)
    except PermissionError as exc:
        raise PermissionError(
            f"'{output_path}' dosyasina yazilamadi. Dosya baska bir programda "
            "(orn. Excel) acik olabilir; kapatip tekrar deneyin."
        ) from exc

    return output_path